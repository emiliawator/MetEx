# xlsx extension
# modified field does not want to be changed :(

from openpyxl import load_workbook
from datetime import datetime

def read(filepath):
    try:
        wb = load_workbook(filepath)
        prop = wb.properties
    except:
        return "Supplied filepath is invalid"

    metadata = []
    
    # str if not date
    metadata.append(tuple(("category", prop.category)))
    metadata.append(tuple(("contentStatus", prop.contentStatus)))
    metadata.append(tuple(("created", prop.created))) #date
    metadata.append(tuple(("creator", prop.creator)))
    metadata.append(tuple(("description", prop.description)))
    metadata.append(tuple(("identifier", prop.identifier)))
    metadata.append(tuple(("keywords", prop.keywords)))
    metadata.append(tuple(("language", prop.language)))
    metadata.append(tuple(("lastModifiedBy", prop.lastModifiedBy)))
    metadata.append(tuple(("lastPrinted", prop.lastPrinted))) #date
    metadata.append(tuple(("modified", prop.modified))) #date zawsze mozna oszukac i zrobic handling ze jezeli obecna data to zmien na hardcoded albo nie mozliwe usuniecie
    metadata.append(tuple(("revision", prop.revision)))
    metadata.append(tuple(("subject", prop.subject)))
    metadata.append(tuple(("title", prop.title)))
    metadata.append(tuple(("version", prop.version)))

    return metadata

def save(metadata, filepath, newfilepath):
    errors = []
    try:
        wb = load_workbook(filepath)
        prop = wb.properties
    except:
        errors.append("Supplied filepath is invalid")
        return errors
    
    notuplelist = [list(i) for i in metadata]

    for item in notuplelist:
        if item[0] == "created" or item[0] == "lastPrinted" or item[0] == "modified": # datetime type
            if item[1] != "None":
                    try:
                        item[1] = datetime.strptime(item[1], "%Y-%m-%d %H:%M:%S")
                    except:
                        errors.append("Correct date format must be supplied in {} field: \n %Y-%m-%d %H:%M:%S".format(item[0]))
                        return errors
            else: # if None by default
                continue
            try:
                setattr(prop, item[0], item[1])
            except:
                errors.append("An error occured while saving {} field".format(item[0]))
                return errors
        else:
            try:
                setattr(prop, item[0], item[1])
            except:
                errors.append("An error occured while saving {} field".format(item[0]))
                return errors
    try:
        wb.save(newfilepath)
    except:
        errors.append("Supplied filepath is invalid")
        return errors
    

def erase(filepath):
    errors = []
    try:
        wb = load_workbook(filepath)
        prop = wb.properties
    except:
        errors.append("Supplied filepath is invalid")
        return errors
    metadata = [] # list of tuples

    # str if not date
    metadata.append(tuple(("category", "")))
    metadata.append(tuple(("contentStatus", "")))
    metadata.append(tuple(("creator", "")))
    metadata.append(tuple(("description", "")))
    metadata.append(tuple(("identifier", "")))
    metadata.append(tuple(("keywords", "")))
    metadata.append(tuple(("language", "")))
    metadata.append(tuple(("lastModifiedBy", "")))
    metadata.append(tuple(("revision", "")))
    metadata.append(tuple(("subject", "")))
    metadata.append(tuple(("title", "")))
    metadata.append(tuple(("version", "")))
    
    # datetime handling
    default_date = datetime.min
    if prop.created is not None:
        metadata.append(tuple(("created", default_date)))
    if prop.lastPrinted is not None:
        metadata.append(tuple(("lastPrinted", default_date)))
    if prop.modified is not None:
        metadata.append(tuple(("modified", default_date)))
    
    notuplelist = [list(i) for i in metadata]

    for item in notuplelist:
        setattr(prop, item[0], item[1])

    try:
        wb.save(filepath)
    except:
        errors.append("Supplied filepath is invalid")
        return errors